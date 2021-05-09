import openpyxl as xl

def working_with_excel(name_of_excel_file,name_of_excel_sheet):
    wb = xl.load_workbook(name_of_excel_file)
    sheet = wb[name_of_excel_sheet]
    return sheet

def find_CpG_Island(name_of_excel_file):
    found = 0
    count = 0
    location = []
    for row in range(2,sheet.max_row+1):
        cell = sheet.cell(row,2)
        nextCell = sheet.cell(row+1,2)
        if cell.value == "c" and nextCell.value =="g":
            count = count + 1
            location.append(row)
    total_number = sheet.max_row-1
    frequency = (count/total_number)*100
    if count>=2:
        found = 1
        print("CpG island found in " + name_of_excel_file + " at cell " + str(location))
    else:
        print("CpG island not found" + name_of_excel_file)

number_of_excel_sheet = int(input("Enter the number of excel sheets "))
for i in range(number_of_excel_sheet):
    name_of_excel_file = "Sheet "+str(i+1)+".xlsx";
    name_of_excel_sheet = "Sheet1";
    sheet = working_with_excel(name_of_excel_file, name_of_excel_sheet)
    find_CpG_Island(name_of_excel_file)